"""
FinHub — Personal Finance Portal
Multi-user edition with per-account data isolation.
Deploy: Render.com (web service, free tier)
"""
import os, io, sqlite3, uuid, hashlib, secrets
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, g, jsonify, request, render_template,
                   send_file, session, redirect, url_for)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Render mounts a persistent disk at /data; fall back to local for dev
DATA_DIR = os.environ.get('DATA_DIR', BASE_DIR)
DB_PATH  = os.path.join(DATA_DIR, 'budget.db')

app = Flask(__name__)
# Use env var SECRET_KEY on Render; fallback for local dev
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Secure cookies on Render (HTTPS); plain HTTP locally
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('RENDER', False)
app.config['SESSION_COOKIE_PATH'] = '/'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# ── Signup control ─────────────────────────────────────────────────────────────
# Set ALLOW_SIGNUP=false in Render env to close public registration
ALLOW_SIGNUP = os.environ.get('ALLOW_SIGNUP', 'true').lower() != 'false'

CATEGORIES = ['income','insurance','investments','expenses','discretionary',
              'credit_card','loan']

# ── DB helpers ────────────────────────────────────────────────────────────────
def get_db():
    db = getattr(g, '_db', None)
    if db is None:
        db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        db.execute('PRAGMA journal_mode=WAL')
        g._db = db
    return db

@app.teardown_appcontext
def close_db(exc):
    db = getattr(g, '_db', None)
    if db: db.close()

def q(sql, params=(), fetchall=False, fetchone=False):
    db  = get_db()
    cur = db.execute(sql, params)
    if fetchall:  return [dict(r) for r in cur.fetchall()]
    if fetchone:
        row = cur.fetchone()
        return dict(row) if row else None
    db.commit()
    return None

def hash_password(pw):
    """PBKDF2-HMAC-SHA256 with random salt."""
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 260000)
    return f'{salt}${h.hex()}'

def verify_password(pw, stored):
    """Verify a password against stored hash."""
    try:
        salt, hx = stored.split('$', 1)
        h = hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 260000)
        return secrets.compare_digest(h.hex(), hx)
    except Exception:
        return False

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id          TEXT PRIMARY KEY,
            username    TEXT NOT NULL UNIQUE,
            email       TEXT NOT NULL DEFAULT '',
            password_hash TEXT NOT NULL,
            is_admin    INTEGER NOT NULL DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS entries (
            id          TEXT PRIMARY KEY,
            user_id     TEXT NOT NULL DEFAULT 'default',
            year        INTEGER NOT NULL,
            month       INTEGER NOT NULL,
            category    TEXT NOT NULL,
            name        TEXT NOT NULL,
            amount      REAL NOT NULL,
            note        TEXT DEFAULT '',
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_ym ON entries(year, month);

        CREATE TABLE IF NOT EXISTS credit_cards (
            id           TEXT PRIMARY KEY,
            user_id      TEXT NOT NULL DEFAULT 'default',
            bank         TEXT NOT NULL,
            last4        TEXT NOT NULL DEFAULT '',
            limit_amt    REAL NOT NULL DEFAULT 0,
            created_at   TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS cc_bills (
            id           TEXT PRIMARY KEY,
            card_id      TEXT NOT NULL REFERENCES credit_cards(id) ON DELETE CASCADE,
            year         INTEGER NOT NULL,
            month        INTEGER NOT NULL,
            due_date     TEXT NOT NULL DEFAULT '',
            total_amt    REAL NOT NULL DEFAULT 0,
            min_due      REAL NOT NULL DEFAULT 0,
            paid_amt     REAL NOT NULL DEFAULT 0,
            status       TEXT NOT NULL DEFAULT 'unpaid',
            note         TEXT DEFAULT '',
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_ccbill_ym ON cc_bills(year, month);

        CREATE TABLE IF NOT EXISTS loans (
            id              TEXT PRIMARY KEY,
            user_id         TEXT NOT NULL DEFAULT 'default',
            lender          TEXT NOT NULL,
            loan_type       TEXT NOT NULL DEFAULT 'personal',
            principal       REAL NOT NULL DEFAULT 0,
            interest_rate   REAL NOT NULL DEFAULT 0,
            tenure_months   INTEGER NOT NULL DEFAULT 12,
            emi_amount      REAL NOT NULL DEFAULT 0,
            start_date      TEXT NOT NULL DEFAULT '',
            outstanding     REAL NOT NULL DEFAULT 0,
            status          TEXT NOT NULL DEFAULT 'active',
            note            TEXT DEFAULT '',
            created_at      TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS loan_payments (
            id          TEXT PRIMARY KEY,
            loan_id     TEXT NOT NULL REFERENCES loans(id) ON DELETE CASCADE,
            year        INTEGER NOT NULL,
            month       INTEGER NOT NULL,
            paid_date   TEXT NOT NULL DEFAULT '',
            emi_amt     REAL NOT NULL DEFAULT 0,
            principal   REAL NOT NULL DEFAULT 0,
            interest    REAL NOT NULL DEFAULT 0,
            note        TEXT DEFAULT '',
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_lpay_ym ON loan_payments(year, month);
    ''')
    # Safe migration: add new columns if missing
    try:
        conn.execute("ALTER TABLE networth_plan ADD COLUMN invest_rows_json TEXT NOT NULL DEFAULT '[]'")
        conn.commit()
    except Exception:
        pass
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS networth_assets (
            id          TEXT PRIMARY KEY,
            user_id     TEXT NOT NULL DEFAULT 'default',
            account     TEXT NOT NULL,
            asset_class TEXT NOT NULL,
            label       TEXT NOT NULL DEFAULT '',
            amount      REAL NOT NULL DEFAULT 0,
            sort_order  INTEGER NOT NULL DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ab_baskets (
            id           TEXT PRIMARY KEY,
            user_id      TEXT NOT NULL DEFAULT 'default',
            name         TEXT NOT NULL,
            strategy     TEXT NOT NULL DEFAULT '',
            rebalance    TEXT NOT NULL DEFAULT 'monthly',
            capital      REAL NOT NULL DEFAULT 0,
            nav          REAL NOT NULL DEFAULT 100.0,
            inception    TEXT NOT NULL DEFAULT (date('now')),
            status       TEXT NOT NULL DEFAULT 'active',
            notes        TEXT NOT NULL DEFAULT '',
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ab_instruments (
            id           TEXT PRIMARY KEY,
            basket_id    TEXT NOT NULL,
            tradingsymbol TEXT NOT NULL,
            exchange     TEXT NOT NULL DEFAULT 'NSE',
            target_pct   REAL NOT NULL DEFAULT 0,
            qty          INTEGER NOT NULL DEFAULT 0,
            avg_price    REAL NOT NULL DEFAULT 0,
            sort_order   INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS ab_nav_history (
            id           TEXT PRIMARY KEY,
            basket_id    TEXT NOT NULL,
            nav_date     TEXT NOT NULL,
            nav          REAL NOT NULL,
            total_value  REAL NOT NULL DEFAULT 0,
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ab_orders (
            id           TEXT PRIMARY KEY,
            basket_id    TEXT NOT NULL,
            order_id     TEXT,
            tradingsymbol TEXT NOT NULL,
            exchange     TEXT NOT NULL DEFAULT 'NSE',
            transaction_type TEXT NOT NULL,
            qty          INTEGER NOT NULL DEFAULT 0,
            price        REAL NOT NULL DEFAULT 0,
            status       TEXT NOT NULL DEFAULT 'pending',
            reason       TEXT NOT NULL DEFAULT '',
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS networth_plan (
            id              TEXT PRIMARY KEY,
            user_id         TEXT NOT NULL DEFAULT 'default',
            current_age     INTEGER NOT NULL DEFAULT 30,
            current_nw      REAL NOT NULL DEFAULT 5000000,
            roi_pct         REAL NOT NULL DEFAULT 15,
            swp_pct         REAL NOT NULL DEFAULT 3,
            swp_start_age   INTEGER NOT NULL DEFAULT 55,
            annual_invest_json TEXT NOT NULL DEFAULT '{}',
            invest_rows_json   TEXT NOT NULL DEFAULT '[]',
            updated_at      TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    # ── Safe migrations for existing databases ────────────────────────────────
    # Add user_id column to tables that need it (silently skip if already exists)
    col_migrations = [
        "ALTER TABLE entries ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
        "ALTER TABLE credit_cards ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
        "ALTER TABLE loans ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
        "ALTER TABLE networth_assets ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
        "ALTER TABLE networth_plan ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
        "ALTER TABLE ab_baskets ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
        "ALTER TABLE kite_snapshots ADD COLUMN user_id TEXT NOT NULL DEFAULT 'default'",
    ]
    for m in col_migrations:
        try: conn.execute(m); conn.commit()
        except Exception: pass  # column already exists

    # kite_config needs special handling: old schema has PRIMARY KEY (key) only.
    # We need PRIMARY KEY (user_id, key). Rebuild the table if needed.
    try:
        cols = [r[1] for r in conn.execute('PRAGMA table_info(kite_config)').fetchall()]
        if 'user_id' not in cols:
            # Old schema — migrate data then rebuild
            conn.executescript('''
                ALTER TABLE kite_config RENAME TO kite_config_old;
                CREATE TABLE kite_config (
                    user_id TEXT NOT NULL DEFAULT 'default',
                    key     TEXT NOT NULL,
                    value   TEXT NOT NULL,
                    PRIMARY KEY (user_id, key)
                );
                INSERT INTO kite_config (user_id, key, value)
                    SELECT 'default', key, value FROM kite_config_old;
                DROP TABLE kite_config_old;
            ''')
            conn.commit()
    except Exception:
        pass  # table doesn't exist yet — init_kite_db will create it

    # ── Seed default admin if no users exist ──────────────────────────────────
    existing = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing == 0:
        import hashlib as _hl, secrets as _sec
        salt = _sec.token_hex(16)
        h = _hl.pbkdf2_hmac('sha256', b'admindr', salt.encode(), 260000)
        conn.execute(
            "INSERT INTO users (id,username,email,password_hash,is_admin) VALUES (?,?,?,?,1)",
            (str(uuid.uuid4()), 'admin', 'admin@finhub.local', f'{salt}${h.hex()}')
        )
        conn.commit()
    conn.close()
    print("✅ Database ready:", DB_PATH)

# ── Auth helpers ──────────────────────────────────────────────────────────────
def current_user_id():
    return session.get('user_id', 'default')

def current_username():
    return session.get('username', 'user')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            wants_json = (
                request.path.startswith('/api/') or
                'application/json' in request.headers.get('Accept', '') or
                request.headers.get('X-Requested-With') == 'XMLHttpRequest' or
                request.headers.get('Content-Type','').startswith('application/json')
            )
            if wants_json:
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def uid():
    """Shorthand for current_user_id() — used in query params.""";
    return current_user_id()

# ── Pages ─────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if not session.get('user_id'):
        return redirect(url_for('login_page'))
    return redirect(url_for('home'))

@app.route('/home')
@login_required
def home():
    return render_template('home.html', username=current_username())

@app.route('/budget')
@login_required
def budget():
    return render_template('index.html')

# ── Login ──────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET'])
def login_page():
    if session.get('user_id'): return redirect(url_for('home'))
    return render_template('login.html', error=None, allow_signup=ALLOW_SIGNUP)

@app.route('/login', methods=['POST'])
def do_login():
    try:
        username = request.form.get('username','').strip().lower()
        password = request.form.get('password','')
        user = q('SELECT * FROM users WHERE LOWER(username)=?', (username,), fetchone=True)
        if user and verify_password(password, user['password_hash']):
            session.permanent = True
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['is_admin']  = bool(user['is_admin'])
            return redirect(url_for('home'))
        return render_template('login.html', error='Invalid username or password.', allow_signup=ALLOW_SIGNUP)
    except Exception as e:
        import traceback as _tb; _tb.print_exc()
        return render_template('login.html', error=f'Login error: {e}', allow_signup=ALLOW_SIGNUP)

# ── Register ───────────────────────────────────────────────────────────────────
@app.route('/register', methods=['GET'])
def register_page():
    if not ALLOW_SIGNUP:
        return redirect(url_for('login_page'))
    if session.get('user_id'): return redirect(url_for('home'))
    return render_template('register.html', error=None)

@app.route('/register', methods=['POST'])
def do_register():
    if not ALLOW_SIGNUP:
        return redirect(url_for('login_page'))
    try:
        username = request.form.get('username','').strip()
        email    = request.form.get('email','').strip().lower()
        password = request.form.get('password','')
        confirm  = request.form.get('confirm','')
        error = None
        if not username or len(username) < 3:
            error = 'Username must be at least 3 characters.'
        elif not password or len(password) < 6:
            error = 'Password must be at least 6 characters.'
        elif password != confirm:
            error = 'Passwords do not match.'
        elif q('SELECT id FROM users WHERE LOWER(username)=?', (username.lower(),), fetchone=True):
            error = 'Username already taken.'
        if error:
            return render_template('register.html', error=error)
        uid_new = str(uuid.uuid4())
        q('INSERT INTO users (id,username,email,password_hash) VALUES (?,?,?,?)',
          (uid_new, username, email, hash_password(password)))
        session.permanent = True
        session['user_id']  = uid_new
        session['username'] = username
        session['is_admin'] = False
        return redirect(url_for('home'))
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        return render_template('register.html',
            error=f'Registration failed: {e}. Please try again or contact admin.')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.errorhandler(500)
def internal_error(e):
    import traceback as _tb
    tb = _tb.format_exc()
    print('500 ERROR:', tb)
    # Return JSON for API calls, HTML for page requests
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error', 'detail': str(e)}), 500
    return f'''<!DOCTYPE html><html><head><title>Error — FinHub</title>
<style>body{{font-family:sans-serif;padding:2rem;background:#f5f0e8}}
pre{{background:#fff;padding:1rem;border-radius:8px;overflow-x:auto;font-size:.85rem;border:1px solid #ddd}}
a{{color:#c9a84c}}</style></head><body>
<h2>⚠️ Something went wrong</h2>
<p>Please <a href="/login">sign in again</a> or <a href="javascript:history.back()">go back</a>.</p>
<details><summary style="cursor:pointer;color:#999;font-size:.85rem">Technical details</summary>
<pre>{str(e)}</pre></details>
</body></html>''', 500

# ── Budget entries API ────────────────────────────────────────────────────────
@app.route('/api/entries', methods=['GET'])
@login_required
def get_entries():
    year  = int(request.args.get('year',  datetime.now().year))
    month = int(request.args.get('month', datetime.now().month - 1))
    rows  = q('SELECT * FROM entries WHERE user_id=? AND year=? AND month=? ORDER BY created_at',
               (uid(), year, month), fetchall=True)
    result = {cat: [] for cat in CATEGORIES}
    for r in rows:
        if r['category'] in result:
            result[r['category']].append({
                'id': r['id'], 'name': r['name'],
                'amount': float(r['amount']), 'note': r['note'] or ''
            })
    return jsonify(result)

@app.route('/api/entries', methods=['POST'])
@login_required
def add_entry():
    d = request.get_json()
    q('INSERT INTO entries (id,user_id,year,month,category,name,amount,note) VALUES (?,?,?,?,?,?,?,?)',
      (d['id'], uid(), d['year'], d['month'], d['category'], d['name'], d['amount'], d.get('note','')))
    return jsonify({'ok': True})

@app.route('/api/entries/<eid>', methods=['PUT'])
@login_required
def update_entry(eid):
    d = request.get_json()
    q('UPDATE entries SET name=?,amount=?,note=? WHERE id=? AND user_id=?',
      (d['name'], d['amount'], d.get('note',''), eid, uid()))
    return jsonify({'ok': True})

@app.route('/api/entries/<eid>', methods=['DELETE'])
@login_required
def delete_entry(eid):
    q('DELETE FROM entries WHERE id=? AND user_id=?', (eid, uid()))
    return jsonify({'ok': True})

# ── Copy previous month ───────────────────────────────────────────────────────
@app.route('/api/entries/copy', methods=['POST'])
@login_required
def copy_prev_month():
    d = request.get_json()
    to_year   = int(d['toYear'])
    to_month  = int(d['toMonth'])
    from_year = int(d['fromYear'])
    from_month= int(d['fromMonth'])
    cats      = d.get('categories', CATEGORIES)   # which cats to copy

    # Check target is not empty (only copy if target has no entries, unless forced)
    force = d.get('force', False)
    if not force:
        existing = q('SELECT COUNT(*) as n FROM entries WHERE user_id=? AND year=? AND month=?',
                     (uid(), to_year, to_month), fetchone=True)
        if existing and existing['n'] > 0:
            return jsonify({'ok': False, 'conflict': True,
                            'count': existing['n']})

    src_rows = q('SELECT * FROM entries WHERE user_id=? AND year=? AND month=? AND category IN ({})'.format(
                  ','.join('?'*len(cats))),
                 (uid(), from_year, from_month, *cats), fetchall=True)

    db = get_db()
    for r in src_rows:
        try:
            db.execute(
                'INSERT INTO entries (id,user_id,year,month,category,name,amount,note) VALUES (?,?,?,?,?,?,?,?)',
                (str(uuid.uuid4()), uid(), to_year, to_month,
                 r['category'], r['name'], r['amount'], r['note'] or '')
            )
        except Exception:
            pass
    db.commit()
    return jsonify({'ok': True, 'copied': len(src_rows)})

# ── Analysis ──────────────────────────────────────────────────────────────────
@app.route('/api/analysis')
@login_required
def get_analysis():
    year = int(request.args.get('year', datetime.now().year))
    rows = q('SELECT month,category,SUM(amount) as total FROM entries WHERE user_id=? AND year=? GROUP BY month,category',
              (uid(), year), fetchall=True)
    data = {m: {cat: 0 for cat in CATEGORIES} for m in range(12)}
    for r in rows:
        if r['month'] in data and r['category'] in data[r['month']]:
            data[r['month']][r['category']] = float(r['total'] or 0)
    return jsonify(data)

@app.route('/api/analysis/years')
@login_required
def get_years():
    rows  = q('SELECT DISTINCT year FROM entries WHERE user_id=? ORDER BY year DESC', (uid(),), fetchall=True)
    years = [r['year'] for r in rows]
    cur   = datetime.now().year
    if cur not in years: years.insert(0, cur)
    return jsonify(years)

# ── Credit Cards API ──────────────────────────────────────────────────────────
@app.route('/api/cards', methods=['GET'])
@login_required
def get_cards():
    return jsonify(q('SELECT * FROM credit_cards WHERE user_id=? ORDER BY bank', (uid(),), fetchall=True))

@app.route('/api/cards', methods=['POST'])
@login_required
def add_card():
    d = request.get_json()
    q('INSERT INTO credit_cards (id,user_id,bank,last4,limit_amt) VALUES (?,?,?,?,?)',
      (str(uuid.uuid4()), uid(), d['bank'], d.get('last4',''), d.get('limit_amt',0)))
    return jsonify({'ok': True})

@app.route('/api/cards/<cid>', methods=['PUT'])
@login_required
def update_card(cid):
    d = request.get_json()
    q('UPDATE credit_cards SET bank=?,last4=?,limit_amt=? WHERE id=?',
      (d['bank'], d.get('last4',''), d.get('limit_amt',0), cid))
    return jsonify({'ok': True})

@app.route('/api/cards/<cid>', methods=['DELETE'])
@login_required
def delete_card(cid):
    q('DELETE FROM credit_cards WHERE id=? AND user_id=?', (cid, uid()))
    return jsonify({'ok': True})

# CC Bills
@app.route('/api/cc-bills', methods=['GET'])
@login_required
def get_cc_bills():
    year  = int(request.args.get('year',  datetime.now().year))
    month = int(request.args.get('month', datetime.now().month - 1))
    rows  = q('''SELECT b.*, c.bank, c.last4, c.limit_amt
                 FROM cc_bills b JOIN credit_cards c ON b.card_id=c.id
                 WHERE c.user_id=? AND b.year=? AND b.month=? ORDER BY c.bank''',
               (uid(), year, month), fetchall=True)
    return jsonify([dict(r) for r in rows])

@app.route('/api/cc-bills', methods=['POST'])
@login_required
def add_cc_bill():
    d = request.get_json()
    q('''INSERT INTO cc_bills (id,card_id,year,month,due_date,total_amt,min_due,paid_amt,status,note)
         VALUES (?,?,?,?,?,?,?,?,?,?)''',
      (str(uuid.uuid4()), d['card_id'], d['year'], d['month'],
       d.get('due_date',''), d.get('total_amt',0), d.get('min_due',0),
       d.get('paid_amt',0), d.get('status','unpaid'), d.get('note','')))
    return jsonify({'ok': True})

@app.route('/api/cc-bills/<bid>', methods=['PUT'])
@login_required
def update_cc_bill(bid):
    d = request.get_json()
    q('''UPDATE cc_bills SET due_date=?,total_amt=?,min_due=?,paid_amt=?,status=?,note=?
         WHERE id=?''',
      (d.get('due_date',''), d.get('total_amt',0), d.get('min_due',0),
       d.get('paid_amt',0), d.get('status','unpaid'), d.get('note',''), bid))
    return jsonify({'ok': True})

@app.route('/api/cc-bills/<bid>', methods=['DELETE'])
@login_required
def delete_cc_bill(bid):
    q('DELETE FROM cc_bills WHERE id=?', (bid,))
    return jsonify({'ok': True})

@app.route('/api/cc-summary')
@login_required
def cc_summary():
    year  = int(request.args.get('year',  datetime.now().year))
    month = int(request.args.get('month', datetime.now().month - 1))
    rows  = q('''SELECT b.*, c.bank, c.last4, c.limit_amt
                 FROM cc_bills b JOIN credit_cards c ON b.card_id=c.id
                 WHERE c.user_id=? AND b.year=? AND b.month=?''', (uid(), year, month), fetchall=True)
    total_due  = sum(float(r['total_amt']) for r in rows)
    total_paid = sum(float(r['paid_amt'])  for r in rows)
    unpaid_cnt = sum(1 for r in rows if r['status'] != 'paid')
    return jsonify({'total_due': total_due, 'total_paid': total_paid,
                    'unpaid': unpaid_cnt, 'bills': len(rows)})

# ── Loans API ─────────────────────────────────────────────────────────────────
@app.route('/api/loans', methods=['GET'])
@login_required
def get_loans():
    status = request.args.get('status', 'all')
    if status == 'active':
        rows = q("SELECT * FROM loans WHERE user_id=? AND status='active' ORDER BY lender", (uid(),), fetchall=True)
    else:
        rows = q('SELECT * FROM loans WHERE user_id=? ORDER BY lender', (uid(),), fetchall=True)
    return jsonify(rows)

@app.route('/api/loans', methods=['POST'])
@login_required
def add_loan():
    d = request.get_json()
    q('''INSERT INTO loans (id,user_id,lender,loan_type,principal,interest_rate,tenure_months,
                            emi_amount,start_date,outstanding,status,note)
         VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
      (str(uuid.uuid4()), uid(), d['lender'], d.get('loan_type','personal'),
       d.get('principal',0), d.get('interest_rate',0), d.get('tenure_months',12),
       d.get('emi_amount',0), d.get('start_date',''),
       d.get('outstanding', d.get('principal',0)),
       d.get('status','active'), d.get('note','')))
    return jsonify({'ok': True})

@app.route('/api/loans/<lid>', methods=['PUT'])
@login_required
def update_loan(lid):
    d = request.get_json()
    q('''UPDATE loans SET lender=?,loan_type=?,principal=?,interest_rate=?,
                          tenure_months=?,emi_amount=?,start_date=?,outstanding=?,status=?,note=?
         WHERE id=? AND user_id=?''',
      (d['lender'], d.get('loan_type','personal'), d.get('principal',0),
       d.get('interest_rate',0), d.get('tenure_months',12), d.get('emi_amount',0),
       d.get('start_date',''), d.get('outstanding',0),
       d.get('status','active'), d.get('note',''), lid, uid()))
    return jsonify({'ok': True})

@app.route('/api/loans/<lid>', methods=['DELETE'])
@login_required
def delete_loan(lid):
    q('DELETE FROM loans WHERE id=? AND user_id=?', (lid, uid()))
    return jsonify({'ok': True})

@app.route('/api/loan-payments', methods=['GET'])
@login_required
def get_loan_payments():
    lid   = request.args.get('loan_id')
    year  = request.args.get('year')
    month = request.args.get('month')
    if lid:
        rows = q('SELECT * FROM loan_payments lp JOIN loans l ON lp.loan_id=l.id WHERE lp.loan_id=? AND l.user_id=? ORDER BY lp.paid_date DESC',
                 (lid, uid()), fetchall=True)
    elif year and month:
        rows = q('''SELECT p.*, l.lender, l.loan_type
                    FROM loan_payments p JOIN loans l ON p.loan_id=l.id
                    WHERE l.user_id=? AND p.year=? AND p.month=? ORDER BY l.lender''',
                 (uid(), int(year), int(month)), fetchall=True)
    else:
        rows = q('SELECT lp.* FROM loan_payments lp JOIN loans l ON lp.loan_id=l.id WHERE l.user_id=? ORDER BY lp.paid_date DESC', (uid(),), fetchall=True)
    return jsonify(rows)

@app.route('/api/loan-payments', methods=['POST'])
@login_required
def add_loan_payment():
    d = request.get_json()
    emi   = float(d.get('emi_amt', 0))
    princ = float(d.get('principal', 0))
    intr  = float(d.get('interest', emi - princ))
    q('''INSERT INTO loan_payments (id,loan_id,year,month,paid_date,emi_amt,principal,interest,note)
         VALUES (?,?,?,?,?,?,?,?,?)''',
      (str(uuid.uuid4()), d['loan_id'], d['year'], d['month'],
       d.get('paid_date',''), emi, princ, intr, d.get('note','')))
    # Update outstanding balance
    if princ > 0:
        q('UPDATE loans SET outstanding = MAX(0, outstanding - ?) WHERE id=?',
          (princ, d['loan_id']))
    return jsonify({'ok': True})

@app.route('/api/loan-payments/<pid>', methods=['DELETE'])
@login_required
def delete_loan_payment(pid):
    # Verify ownership via parent loan
    pay = q('SELECT lp.* FROM loan_payments lp JOIN loans l ON lp.loan_id=l.id WHERE lp.id=? AND l.user_id=?', (pid, uid()), fetchone=True)
    if pay:
        q('UPDATE loans SET outstanding = outstanding + ? WHERE id=? AND user_id=?',
          (float(pay['principal']), pay['loan_id'], uid()))
        q('DELETE FROM loan_payments WHERE id=?', (pid,))
    return jsonify({'ok': True})

@app.route('/api/loan-summary')
@login_required
def loan_summary():
    loans = q("SELECT * FROM loans WHERE user_id=? AND status='active'", (uid(),), fetchall=True)
    total_outstanding = sum(float(l['outstanding']) for l in loans)
    total_emi         = sum(float(l['emi_amount'])  for l in loans)
    return jsonify({'total_outstanding': total_outstanding,
                    'total_emi': total_emi, 'active_loans': len(loans)})

# ── Excel exports ─────────────────────────────────────────────────────────────
@app.route('/api/export/month')
@login_required
def export_month():
    from openpyxl import Workbook
    year  = int(request.args.get('year',  datetime.now().year))
    month = int(request.args.get('month', datetime.now().month - 1))
    MN = ['January','February','March','April','May','June',
          'July','August','September','October','November','December']
    lbl   = f"{MN[month]} {year}"
    rows  = q('SELECT * FROM entries WHERE user_id=? AND year=? AND month=? ORDER BY category,created_at',
               (uid(), year, month), fetchall=True)
    by_cat= {cat: [r for r in rows if r['category']==cat] for cat in CATEGORIES}
    wb    = Workbook(); ws = wb.active; ws.title = 'Summary'
    _summary_sheet(ws, lbl, by_cat)
    CC = {'income':'2D6A4F','insurance':'1D3557','investments':'7B2D8B','expenses':'9C2323',
          'discretionary':'B5591A','credit_card':'0D47A1','loan':'4A0E0E'}
    CL = {'income':'Income','insurance':'Insurance','investments':'Investments',
          'expenses':'Expenses','discretionary':'Discretionary',
          'credit_card':'Credit Card','loan':'Loan'}
    for cat in CATEGORIES:
        _cat_sheet(wb.create_sheet(CL.get(cat,cat)), CL.get(cat,cat), lbl, by_cat[cat], CC.get(cat,'888888'))
    # CC bills sheet
    bills = q('''SELECT b.*,c.bank,c.last4 FROM cc_bills b JOIN credit_cards c ON b.card_id=c.id
                 WHERE c.user_id=? AND b.year=? AND b.month=? ORDER BY c.bank''', (uid(),year,month), fetchall=True)
    if bills:
        wcc = wb.create_sheet('CC Bills Detail')
        _cc_bills_sheet(wcc, lbl, bills)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"BudgetCraft_{MN[month]}_{year}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export/full')
@login_required
def export_full():
    from openpyxl import Workbook
    year = int(request.args.get('year', datetime.now().year))
    MN = ['January','February','March','April','May','June',
          'July','August','September','October','November','December']
    MS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    oc = ['insurance','investments','expenses','discretionary','credit_card','loan']
    ac = ['income'] + oc
    ry = q('SELECT * FROM entries WHERE user_id=? AND year=? ORDER BY month,category,created_at', (uid(),year), fetchall=True)
    ra = q('SELECT * FROM entries WHERE user_id=? ORDER BY year,month,category', (uid(),), fetchall=True)
    wb = Workbook(); ws = wb.active; ws.title = 'Monthly Summary'
    _period_sheet(ws, f'Monthly Summary — {year}', MN, ry, ac, oc, 'monthly')
    _period_sheet(wb.create_sheet('Quarterly Summary'), f'Quarterly Summary — {year}',
                  ['Q1 (Jan-Mar)','Q2 (Apr-Jun)','Q3 (Jul-Sep)','Q4 (Oct-Dec)'], ry, ac, oc, 'quarterly')
    yrs = sorted(set(r['year'] for r in ra))
    if year not in yrs: yrs.append(year)
    _period_sheet(wb.create_sheet('Year-over-Year'), 'Year-over-Year',
                  [str(y) for y in sorted(yrs)], ra, ac, oc, 'yearly', years=sorted(yrs))
    for m in range(12):
        _month_detail(wb.create_sheet(f"{MS[m]} {year}"), f"{MN[m]} {year}",
                      [r for r in ry if r['month']==m], ac)
    # Loans sheet
    loans = q('SELECT * FROM loans WHERE user_id=? ORDER BY lender', (uid(),), fetchall=True)
    if loans:
        wl = wb.create_sheet('Loans')
        _loans_sheet(wl, loans)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"BudgetCraft_FullSummary_{year}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Excel helpers ─────────────────────────────────────────────────────────────
def _cell(ws, r, c, val, bold=False, bg=None, fg='000000', sz=10, align='left', num_fmt=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, color=fg, size=sz, name='Arial')
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if bg:   cell.fill = PatternFill('solid', start_color=bg, end_color=bg)
    if num_fmt: cell.number_format = num_fmt
    return cell

def _num(ws, r, c, val):
    return _cell(ws, r, c, round(float(val or 0), 2), align='right', num_fmt='#,##0.00')

def _summary_sheet(ws, lbl, by_cat):
    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=12; ws.column_dimensions['C'].width=18
    _cell(ws,1,1,f'BudgetCraft — {lbl}',bold=True,bg='0F0E0D',fg='F5F0E8',sz=13)
    ws.merge_cells('A1:C1'); ws.row_dimensions[1].height=28
    for ci,(h,al) in enumerate([('Category','left'),('Entries','right'),('Total','right')],1):
        _cell(ws,3,ci,h,bold=True,bg='EDE7D9',align=al)
    CL={'income':'Income','insurance':'Insurance','investments':'Investments','expenses':'Expenses',
        'discretionary':'Discretionary','credit_card':'Credit Card','loan':'Loan'}
    CC={'income':'2D6A4F','insurance':'1D3557','investments':'7B2D8B','expenses':'9C2323',
        'discretionary':'B5591A','credit_card':'0D47A1','loan':'4A0E0E'}
    r=4; tI=tO=0
    for cat in CATEGORIES:
        e=by_cat.get(cat,[]); tot=sum(float(x['amount']) for x in e)
        _cell(ws,r,1,CL.get(cat,cat),bold=True,fg=CC.get(cat,'000000'))
        _cell(ws,r,2,len(e),align='right'); _num(ws,r,3,tot)
        if cat=='income': tI=tot
        else: tO+=tot
        r+=1
    r+=1
    for lb,v in [('Total Income',tI),('Total Outflow',tO),('Net Balance',tI-tO)]:
        _cell(ws,r,1,lb,bold=True,bg='F5F0E8'); _num(ws,r,3,v); r+=1

def _cat_sheet(ws, cat_label, lbl, entries, color):
    for col,w in zip('ABCD',[5,30,24,16]): ws.column_dimensions[col].width=w
    _cell(ws,1,1,f'{cat_label} — {lbl}',bold=True,bg=color,fg='FFFFFF',sz=12); ws.merge_cells('A1:D1')
    for ci,(h,al) in enumerate([('#','center'),('Name','left'),('Note','left'),('Amount','right')],1):
        _cell(ws,3,ci,h,bold=True,bg='EDE7D9',align=al)
    for i,e in enumerate(entries,1):
        ws.cell(row=3+i,column=1,value=i); ws.cell(row=3+i,column=2,value=e['name'])
        ws.cell(row=3+i,column=3,value=e.get('note','')); _num(ws,3+i,4,e['amount'])
    r=3+len(entries)+1; _cell(ws,r,3,'TOTAL',bold=True,bg='EDE7D9')
    _num(ws,r,4,sum(float(e['amount']) for e in entries))

def _cc_bills_sheet(ws, lbl, bills):
    ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=10
    ws.column_dimensions['C'].width=14; ws.column_dimensions['D'].width=14
    ws.column_dimensions['E'].width=14; ws.column_dimensions['F'].width=12
    ws.column_dimensions['G'].width=16
    _cell(ws,1,1,f'Credit Card Bills — {lbl}',bold=True,bg='0D47A1',fg='FFFFFF',sz=12)
    ws.merge_cells('A1:G1')
    for ci,h in enumerate(['Bank / Card','Last 4','Due Date','Total Amt','Min Due','Paid','Status'],1):
        _cell(ws,3,ci,h,bold=True,bg='EDE7D9',align='right' if ci>2 else 'left')
    for i,b in enumerate(bills,4):
        ws.cell(row=i,column=1,value=b['bank']); ws.cell(row=i,column=2,value=b.get('last4',''))
        ws.cell(row=i,column=3,value=b.get('due_date',''))
        _num(ws,i,4,b['total_amt']); _num(ws,i,5,b['min_due']); _num(ws,i,6,b['paid_amt'])
        ws.cell(row=i,column=7,value=b.get('status','').upper())

def _loans_sheet(ws, loans):
    cols=['Lender','Type','Principal','Rate %','Tenure','EMI','Outstanding','Status','Note']
    ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=14
    for i,c in enumerate('CDEFGHI',1): ws.column_dimensions[c].width=14
    _cell(ws,1,1,'Loans Overview',bold=True,bg='4A0E0E',fg='FFFFFF',sz=12); ws.merge_cells('A1:I1')
    for ci,h in enumerate(cols,1): _cell(ws,3,ci,h,bold=True,bg='EDE7D9',align='right' if ci>2 else 'left')
    for i,l in enumerate(loans,4):
        ws.cell(row=i,column=1,value=l['lender']); ws.cell(row=i,column=2,value=l['loan_type'])
        _num(ws,i,3,l['principal']); ws.cell(row=i,column=4,value=l['interest_rate'])
        ws.cell(row=i,column=5,value=l['tenure_months'])
        _num(ws,i,6,l['emi_amount']); _num(ws,i,7,l['outstanding'])
        ws.cell(row=i,column=8,value=l['status']); ws.cell(row=i,column=9,value=l.get('note',''))

def _period_sheet(ws, title, labels, rows, ac, oc, mode, years=None):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions['A'].width=18
    for i in range(2,2+len(ac)+3): ws.column_dimensions[get_column_letter(i)].width=16
    _cell(ws,1,1,title,bold=True,bg='0F0E0D',fg='F5F0E8',sz=13)
    ws.merge_cells(f'A1:{get_column_letter(1+len(ac)+2)}1')
    CL={'income':'Income','insurance':'Insurance','investments':'Investments','expenses':'Expenses',
        'discretionary':'Discretionary','credit_card':'Credit Card','loan':'Loan'}
    hdr=['Period']+[CL.get(c,c) for c in ac]+['Total Outflow','Net Balance','Savings %']
    for ci,h in enumerate(hdr,1): _cell(ws,3,ci,h,bold=True,bg='EDE7D9',align='right' if ci>1 else 'left')
    def agg(rl):
        d={c:sum(float(r['amount']) for r in rl if r['category']==c) for c in ac}
        d['outflow']=sum(d[c] for c in oc); d['net']=d['income']-d['outflow']; return d
    if   mode=='monthly':   pd2=[(lb,agg([r for r in rows if r['month']==i])) for i,lb in enumerate(labels)]
    elif mode=='quarterly': pd2=[(labels[q],agg([r for r in rows if r['month'] in [q*3,q*3+1,q*3+2]])) for q in range(4)]
    else:                   pd2=[(str(y),agg([r for r in rows if r['year']==y])) for y in (years or [])]
    tI=tO=0; tC={c:0 for c in ac}
    for ri,(lb,d) in enumerate(pd2,4):
        _cell(ws,ri,1,lb,bold=True); rt=round(d['net']/d['income']*100,1) if d['income'] else 0
        for ci,c in enumerate(ac,2): _num(ws,ri,ci,d[c]); tC[c]+=d[c]
        _num(ws,ri,2+len(ac),d['outflow']); _num(ws,ri,2+len(ac)+1,d['net'])
        _cell(ws,ri,2+len(ac)+2,f'{rt}%',align='right'); tI+=d['income']; tO+=d['outflow']
    tr=4+len(pd2); _cell(ws,tr,1,'TOTAL',bold=True,bg='F5F0E8')
    for ci,c in enumerate(ac,2): _num(ws,tr,ci,tC[c])
    _num(ws,tr,2+len(ac),tO); _num(ws,tr,2+len(ac)+1,tI-tO)
    _cell(ws,tr,2+len(ac)+2,f'{round((tI-tO)/tI*100,1) if tI else 0}%',bold=True,align='right')

def _month_detail(ws, lbl, rows, ac):
    for col,w in zip('ABCD',[22,28,20,16]): ws.column_dimensions[col].width=w
    _cell(ws,1,1,f'Detail: {lbl}',bold=True,bg='0F0E0D',fg='F5F0E8',sz=12); ws.merge_cells('A1:D1')
    CL={'income':'Income','insurance':'Insurance','investments':'Investments','expenses':'Expenses',
        'discretionary':'Discretionary','credit_card':'Credit Card','loan':'Loan'}
    r=3
    for cat in ac:
        cr=[x for x in rows if x['category']==cat]
        if not cr: continue
        _cell(ws,r,1,CL.get(cat,cat),bold=True,bg='EDE7D9'); ws.merge_cells(f'A{r}:D{r}'); r+=1
        for e in cr:
            ws.cell(row=r,column=2,value=e['name']); ws.cell(row=r,column=3,value=e.get('note',''))
            _num(ws,r,4,e['amount']); r+=1
        _cell(ws,r,3,'Subtotal',bold=True,bg='F5F0E8')
        _num(ws,r,4,sum(float(e['amount']) for e in cr)); r+=2


# ══════════════════════════════════════════════════════════════════════════════
# INVESTMENT TRACKER — Kite Connect Integration
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/investments')
@login_required
def investments():
    return render_template('investments.html')

@app.route('/strategies')
@login_required
def strategies():
    return render_template('strategies.html')

@app.route('/networth')
@login_required
def networth():
    return render_template('networth.html')

@app.route('/autobasket')
@login_required
def autobasket():
    return render_template('autobasket.html')

def init_kite_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS kite_config (
            user_id TEXT NOT NULL DEFAULT 'default',
            key     TEXT NOT NULL,
            value   TEXT NOT NULL,
            PRIMARY KEY (user_id, key)
        );
        CREATE TABLE IF NOT EXISTS networth_assets (
            id          TEXT PRIMARY KEY,
            user_id     TEXT NOT NULL DEFAULT 'default',
            account     TEXT NOT NULL,
            asset_class TEXT NOT NULL,
            label       TEXT NOT NULL DEFAULT '',
            amount      REAL NOT NULL DEFAULT 0,
            sort_order  INTEGER NOT NULL DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ab_baskets (
            id           TEXT PRIMARY KEY,
            user_id      TEXT NOT NULL DEFAULT 'default',
            name         TEXT NOT NULL,
            strategy     TEXT NOT NULL DEFAULT '',
            rebalance    TEXT NOT NULL DEFAULT 'monthly',
            capital      REAL NOT NULL DEFAULT 0,
            nav          REAL NOT NULL DEFAULT 100.0,
            inception    TEXT NOT NULL DEFAULT (date('now')),
            status       TEXT NOT NULL DEFAULT 'active',
            notes        TEXT NOT NULL DEFAULT '',
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ab_instruments (
            id           TEXT PRIMARY KEY,
            basket_id    TEXT NOT NULL,
            tradingsymbol TEXT NOT NULL,
            exchange     TEXT NOT NULL DEFAULT 'NSE',
            target_pct   REAL NOT NULL DEFAULT 0,
            qty          INTEGER NOT NULL DEFAULT 0,
            avg_price    REAL NOT NULL DEFAULT 0,
            sort_order   INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS ab_nav_history (
            id           TEXT PRIMARY KEY,
            basket_id    TEXT NOT NULL,
            nav_date     TEXT NOT NULL,
            nav          REAL NOT NULL,
            total_value  REAL NOT NULL DEFAULT 0,
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ab_orders (
            id           TEXT PRIMARY KEY,
            basket_id    TEXT NOT NULL,
            order_id     TEXT,
            tradingsymbol TEXT NOT NULL,
            exchange     TEXT NOT NULL DEFAULT 'NSE',
            transaction_type TEXT NOT NULL,
            qty          INTEGER NOT NULL DEFAULT 0,
            price        REAL NOT NULL DEFAULT 0,
            status       TEXT NOT NULL DEFAULT 'pending',
            reason       TEXT NOT NULL DEFAULT '',
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS networth_plan (
            id              TEXT PRIMARY KEY,
            user_id         TEXT NOT NULL DEFAULT 'default',
            current_age     INTEGER NOT NULL DEFAULT 30,
            current_nw      REAL NOT NULL DEFAULT 5000000,
            roi_pct         REAL NOT NULL DEFAULT 15,
            swp_pct         REAL NOT NULL DEFAULT 3,
            swp_start_age   INTEGER NOT NULL DEFAULT 55,
            annual_invest_json TEXT NOT NULL DEFAULT '{}',
            invest_rows_json   TEXT NOT NULL DEFAULT '[]',
            updated_at      TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS kite_snapshots (
            id            TEXT PRIMARY KEY,
            user_id       TEXT NOT NULL DEFAULT 'default',
            snapshot_date TEXT NOT NULL,
            holdings_json TEXT NOT NULL,
            mf_json       TEXT NOT NULL DEFAULT '[]',
            total_value   REAL NOT NULL DEFAULT 0,
            total_pnl     REAL NOT NULL DEFAULT 0,
            created_at    TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    conn.close()

def kite_cfg_get(key):
    row = q('SELECT value FROM kite_config WHERE user_id=? AND key=?', (uid(),key), fetchone=True)
    return row['value'] if row else None

def kite_cfg_set(key, value):
    q('INSERT OR REPLACE INTO kite_config (user_id,key,value) VALUES (?,?,?)', (uid(),key,value))

def kite_api(path, access_token, api_key):
    import urllib.request, urllib.error, json as _json
    url = f'https://api.kite.trade{path}'
    req = urllib.request.Request(url)
    req.add_header('X-Kite-Version', '3')
    req.add_header('Authorization', f'token {api_key}:{access_token}')
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return _json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        # Read the actual Kite error response body
        try:
            body = _json.loads(e.read().decode())
            msg = body.get('message') or body.get('error') or f'HTTP {e.code}'
        except Exception:
            msg = f'HTTP {e.code}: {e.reason}'
        return {'status': 'error', 'message': msg, 'http_code': e.code}
    except urllib.error.URLError as e:
        return {'status': 'error', 'message': f'Network error: {e.reason}'}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}

@app.route('/api/kite/config', methods=['GET'])
@login_required
def kite_config_get():
    return jsonify({
        'api_key':    kite_cfg_get('api_key') or '',
        'has_secret': bool(kite_cfg_get('api_secret')),
        'has_token':  bool(kite_cfg_get('access_token')),
        'redirect_url': request.host_url.rstrip('/') + '/api/kite/callback'
    })

@app.route('/api/kite/config', methods=['POST'])
@login_required
def kite_config_save():
    d = request.get_json()
    if d.get('api_key'):      kite_cfg_set('api_key',      d['api_key'].strip())
    if d.get('api_secret'):   kite_cfg_set('api_secret',   d['api_secret'].strip())
    if d.get('access_token'): kite_cfg_set('access_token', d['access_token'].strip())
    if d.get('user_id'):      kite_cfg_set('user_id',      d['user_id'].strip())
    if d.get('password'):     kite_cfg_set('password',     d['password'])
    if d.get('totp_secret'):  kite_cfg_set('totp_secret',  d['totp_secret'].strip())
    return jsonify({'ok': True})

@app.route('/api/kite/login-url')
@login_required
def kite_login_url():
    api_key = kite_cfg_get('api_key')
    if not api_key:
        return jsonify({'error': 'API key not configured'}), 400
    url = f'https://kite.zerodha.com/connect/login?api_key={api_key}&v=3'
    redirect_url = request.host_url.rstrip('/') + '/api/kite/callback'
    return jsonify({'url': url, 'redirect_url': redirect_url})

@app.route('/api/kite/callback')
@login_required
def kite_callback():
    request_token = request.args.get('request_token', '')
    if not request_token:
        return redirect('/investments?error=no_token')
    api_key    = kite_cfg_get('api_key')
    api_secret = kite_cfg_get('api_secret')
    if not api_key or not api_secret:
        return redirect('/investments?error=no_credentials')
    try:
        import hashlib, urllib.request, urllib.parse, json as _json
        checksum = hashlib.sha256(f'{api_key}{request_token}{api_secret}'.encode()).hexdigest()
        data = urllib.parse.urlencode({
            'api_key': api_key, 'request_token': request_token, 'checksum': checksum
        }).encode()
        req = urllib.request.Request('https://api.kite.trade/session/token',
                                      data=data, method='POST')
        req.add_header('X-Kite-Version', '3')
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = _json.loads(resp.read().decode())
        if result.get('status') == 'success':
            kite_cfg_set('access_token', result['data']['access_token'])
            return redirect('/investments?connected=1')
        return redirect(f'/investments?error=auth_failed')
    except Exception as e:
        return redirect(f'/investments?error=exception')

@app.route('/api/kite/disconnect', methods=['POST'])
@login_required
def kite_disconnect():
    q("DELETE FROM kite_config WHERE user_id=? AND key='access_token'", (uid(),))
    return jsonify({'ok': True})

@app.route('/api/kite/generate-token', methods=['POST'])
@login_required
def kite_generate_token():
    """Exchange a request_token for an access_token manually."""
    import hashlib, urllib.request, urllib.parse, json as _json
    d = request.get_json()
    request_token = (d.get('request_token') or '').strip()
    api_key    = kite_cfg_get('api_key')
    api_secret = kite_cfg_get('api_secret')
    if not request_token:
        return jsonify({'error': 'request_token is required'}), 400
    if not api_key or not api_secret:
        return jsonify({'error': 'API key and secret must be saved first'}), 400
    try:
        checksum = hashlib.sha256(f'{api_key}{request_token}{api_secret}'.encode()).hexdigest()
        data = urllib.parse.urlencode({
            'api_key': api_key, 'request_token': request_token, 'checksum': checksum
        }).encode()
        req = urllib.request.Request('https://api.kite.trade/session/token',
                                      data=data, method='POST')
        req.add_header('X-Kite-Version', '3')
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = _json.loads(resp.read().decode())
        if result.get('status') == 'success':
            token = result['data']['access_token']
            kite_cfg_set('access_token', token)
            return jsonify({'ok': True, 'access_token': token[:8]+'…'})
        return jsonify({'error': result.get('message', 'Token exchange failed')}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/kite/fetch', methods=['POST'])
@login_required
def kite_fetch():
    import json as _json, uuid as _uuid
    access_token = kite_cfg_get('access_token')
    api_key      = kite_cfg_get('api_key')
    if not access_token or not api_key:
        return jsonify({'error': 'Not connected to Kite. Run kite_auto_login.py first.'}), 400

    # ── Fetch equity holdings (unpledged + pledged) ────────────────────────────
    hold_resp = kite_api('/portfolio/holdings', access_token, api_key)
    if hold_resp.get('status') != 'success':
        http_code = hold_resp.get('http_code', 0)
        msg = hold_resp.get('message', 'Holdings fetch failed')
        if http_code == 403 or 'token' in msg.lower() or 'invalid' in msg.lower():
            msg = f'Access token expired or invalid. Re-run kite_auto_login.py. (Kite: {msg})'
        elif http_code == 429:
            msg = 'Rate limit hit. Wait a minute and try again.'
        return jsonify({'error': msg}), 400

    raw_holdings = hold_resp.get('data', [])

    # Augment each holding: add total_quantity (free + pledged) and pledged flag
    holdings = []
    for h in raw_holdings:
        h = dict(h)
        free_qty     = int(h.get('quantity', 0))
        pledged_qty  = int(h.get('collateral_quantity', 0))
        t1_qty       = int(h.get('t1_quantity', 0))
        total_qty    = free_qty + pledged_qty + t1_qty
        h['free_quantity']    = free_qty
        h['pledged_quantity'] = pledged_qty
        h['t1_quantity']      = t1_qty
        h['total_quantity']   = total_qty         # full DEMAT position
        h['is_pledged']       = pledged_qty > 0
        # Recompute value and pnl using total_quantity
        ltp  = float(h.get('last_price', 0))
        avg  = float(h.get('average_price', 0))
        h['current_value']  = ltp * total_qty
        h['invested_value'] = avg * total_qty
        h['total_pnl']      = (ltp - avg) * total_qty
        h['pnl_pct']        = ((ltp - avg) / avg * 100) if avg else 0
        holdings.append(h)

    # ── Fetch MF holdings from the correct endpoint ────────────────────────────
    mf_resp = kite_api('/mf/holdings', access_token, api_key)
    mf_data = mf_resp.get('data', []) if mf_resp.get('status') == 'success' else []

    # Augment MF data for display consistency
    mf_holdings = []
    for m in mf_data:
        m = dict(m)
        qty    = float(m.get('quantity', 0))
        nav    = float(m.get('last_price', 0))
        avg_nav = float(m.get('average_price', 0))
        m['current_value']  = nav * qty
        m['invested_value'] = avg_nav * qty
        m['total_pnl']      = (nav - avg_nav) * qty
        m['pnl_pct']        = ((nav - avg_nav) / avg_nav * 100) if avg_nav else 0
        m['total_quantity'] = qty
        mf_holdings.append(m)

    today = datetime.now().strftime('%Y-%m-%d')

    # Totals include pledged stocks + MF
    eq_value  = sum(h['current_value'] for h in holdings)
    mf_value  = sum(m['current_value'] for m in mf_holdings)
    total_value = eq_value + mf_value
    total_pnl   = sum(h['total_pnl'] for h in holdings) + sum(m['total_pnl'] for m in mf_holdings)

    # Count pledged
    pledged_count = sum(1 for h in holdings if h['is_pledged'])

    # Delete existing snapshot for today (keep only latest per day)
    q("DELETE FROM kite_snapshots WHERE user_id=? AND snapshot_date=?", (uid(), today))
    snap_id = str(_uuid.uuid4())
    q("""INSERT INTO kite_snapshots (id,user_id,snapshot_date,holdings_json,mf_json,total_value,total_pnl)
         VALUES (?,?,?,?,?,?,?)""",
      (snap_id, uid(), today, _json.dumps(holdings), _json.dumps(mf_holdings), total_value, total_pnl))

    return jsonify({
        'ok': True,
        'holdings':      len(holdings),
        'pledged':       pledged_count,
        'mf':            len(mf_holdings),
        'eq_value':      eq_value,
        'mf_value':      mf_value,
        'total_value':   total_value,
        'total_pnl':     total_pnl,
    })

@app.route('/api/kite/holdings')
@login_required
def kite_holdings():
    import json as _json
    snap = q('SELECT * FROM kite_snapshots WHERE user_id=? ORDER BY created_at DESC LIMIT 1', (uid(),), fetchone=True)
    if not snap:
        return jsonify({'holdings': [], 'mf': [], 'fetched_at': None,
                        'total_value': 0, 'total_pnl': 0})

    raw_holdings = _json.loads(snap['holdings_json'])
    raw_mf       = _json.loads(snap['mf_json'])

    # Re-augment equity holdings — ensure computed fields exist even on old snapshots
    holdings_out = []
    for h in raw_holdings:
        h = dict(h)
        free_qty    = int(h.get('free_quantity',    h.get('quantity', 0)))
        pledged_qty = int(h.get('pledged_quantity', h.get('collateral_quantity', 0)))
        t1_qty      = int(h.get('t1_quantity',      0))
        total_qty   = int(h.get('total_quantity',   free_qty + pledged_qty + t1_qty))
        ltp = float(h.get('last_price', 0))
        avg = float(h.get('average_price', 0))
        h['free_quantity']    = free_qty
        h['pledged_quantity'] = pledged_qty
        h['t1_quantity']      = t1_qty
        h['total_quantity']   = total_qty
        h['is_pledged']       = pledged_qty > 0
        h['current_value']    = ltp * total_qty
        h['invested_value']   = avg * total_qty
        h['total_pnl']        = (ltp - avg) * total_qty
        h['pnl_pct']          = ((ltp - avg) / avg * 100) if avg else 0
        holdings_out.append(h)

    # Re-augment MF holdings — ensure computed fields exist on old snapshots
    mf_out = []
    for m in raw_mf:
        m = dict(m)
        qty     = float(m.get('total_quantity', m.get('quantity', 0)))
        nav     = float(m.get('last_price', 0))
        avg_nav = float(m.get('average_price', 0))
        m['total_quantity']  = qty
        m['current_value']   = float(m.get('current_value',  nav * qty))
        m['invested_value']  = float(m.get('invested_value', avg_nav * qty))
        m['total_pnl']       = float(m.get('total_pnl',      (nav - avg_nav) * qty))
        m['pnl_pct']         = float(m.get('pnl_pct',        ((nav - avg_nav) / avg_nav * 100) if avg_nav else 0))
        mf_out.append(m)

    # Recompute totals from augmented data
    total_value = sum(h['current_value'] for h in holdings_out) + sum(m['current_value'] for m in mf_out)
    total_pnl   = sum(h['total_pnl']     for h in holdings_out) + sum(m['total_pnl']     for m in mf_out)

    return jsonify({
        'holdings':      holdings_out,
        'mf':            mf_out,
        'fetched_at':    snap['created_at'],
        'snapshot_date': snap['snapshot_date'],
        'total_value':   total_value,
        'total_pnl':     total_pnl,
    })

@app.route('/api/kite/snapshots')
@login_required
def kite_snapshots():
    # Return one row per calendar date (latest sync of that day)
    rows = q("""
        SELECT snapshot_date, total_value, total_pnl, MAX(created_at) as created_at
        FROM kite_snapshots WHERE user_id=?
        GROUP BY snapshot_date
        ORDER BY snapshot_date ASC
    """, (uid(),), fetchall=True)
    return jsonify(rows)

@app.route('/api/kite/snapshots', methods=['DELETE'])
@login_required
def kite_clear_snapshots():
    q('DELETE FROM kite_snapshots WHERE user_id=?', (uid(),))
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════════════════
# KITE AUTO-LOGIN (embedded — no external script needed)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/kite/auto-login', methods=['POST'])
@login_required
def kite_auto_login():
    """Run TOTP auto-login inline and save access_token to DB."""
    try:
        import hashlib, requests as _req, pyotp as _pyotp
        from urllib.parse import urlparse, parse_qs
    except ImportError as e:
        missing = str(e).split("'")[1] if "'" in str(e) else str(e)
        return jsonify({'error': f"Missing package: {missing}. Run: pip install requests pyotp"}), 400

    api_key    = kite_cfg_get('api_key')
    api_secret = kite_cfg_get('api_secret')
    user_id    = kite_cfg_get('user_id')
    password   = kite_cfg_get('password')
    totp_secret= kite_cfg_get('totp_secret')

    missing = [k for k,v in [('API Key',api_key),('API Secret',api_secret),
                               ('Zerodha User ID',user_id),('Password',password),
                               ('TOTP Secret',totp_secret)] if not v]
    if missing:
        return jsonify({'error': f"Missing credentials: {', '.join(missing)}. Open ⚙ Kite Settings and save them first."}), 400

    steps = []
    try:
        sess = _req.Session()
        sess.headers.update({'X-Kite-Version': '3', 'User-Agent': 'Mozilla/5.0'})

        # Step 1 — Password login
        r = sess.post('https://kite.zerodha.com/api/login',
                      data={'user_id': user_id, 'password': password}, timeout=15)
        r.raise_for_status()
        d = r.json()
        if d.get('status') != 'success':
            return jsonify({'error': f"Login failed: {d.get('message','Wrong credentials')}"}), 400
        req_id = d['data']['request_id']
        steps.append('✓ Password accepted')

        # Step 2 — TOTP 2FA
        totp_val = _pyotp.TOTP(totp_secret).now()
        r = sess.post('https://kite.zerodha.com/api/twofa',
                      data={'user_id': user_id, 'request_id': req_id,
                            'twofa_value': totp_val, 'skip_session': ''}, timeout=15)
        r.raise_for_status()
        d = r.json()
        if d.get('status') != 'success':
            return jsonify({'error': f"2FA failed: {d.get('message','Check TOTP secret')}"}), 400
        steps.append('✓ TOTP accepted')

        # Step 3 — Get request_token from redirect
        r = sess.get(f'https://kite.zerodha.com/connect/login?api_key={api_key}&v=3',
                     allow_redirects=False, timeout=15)
        location = r.headers.get('Location', '')
        req_token = None
        for part in location.split('&'):
            if 'request_token=' in part:
                req_token = part.split('request_token=')[-1].split('&')[0]
                break
        if not req_token:
            r2 = sess.get(f'https://kite.zerodha.com/connect/login?api_key={api_key}&v=3', timeout=15)
            params = parse_qs(urlparse(r2.url).query)
            req_token = (params.get('request_token') or [''])[0]
        if not req_token:
            return jsonify({'error': 'Could not get request_token. Check API Key is correct.'}), 400
        steps.append('✓ request_token obtained')

        # Step 4 — Exchange for access_token
        checksum = hashlib.sha256(f'{api_key}{req_token}{api_secret}'.encode()).hexdigest()
        r = sess.post('https://api.kite.trade/session/token',
                      data={'api_key': api_key, 'request_token': req_token,
                            'checksum': checksum}, timeout=15)
        r.raise_for_status()
        d = r.json()
        if d.get('status') != 'success':
            return jsonify({'error': f"Token exchange failed: {d.get('message','')}"}), 400
        access_token = d['data']['access_token']
        steps.append('✓ access_token received')

        # Save to DB
        kite_cfg_set('access_token', access_token)
        steps.append('✓ Token saved to database')

        return jsonify({'ok': True, 'steps': steps,
                        'token_preview': access_token[:12] + '…'})

    except _req.exceptions.ConnectionError:
        return jsonify({'error': 'Network error — check internet connection'}), 400
    except _req.exceptions.Timeout:
        return jsonify({'error': 'Request timed out — Zerodha servers may be slow'}), 400
    except Exception as e:
        return jsonify({'error': str(e), 'steps': steps}), 400

# ══════════════════════════════════════════════════════════════════════════════
# NET WORTH MODULE API
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/nw/assets', methods=['GET'])
@login_required
def nw_get_assets():
    rows = q('SELECT * FROM networth_assets WHERE user_id=? ORDER BY sort_order, created_at', (uid(),), fetchall=True)
    return jsonify(rows)

@app.route('/api/nw/assets', methods=['POST'])
@login_required
def nw_add_asset():
    import uuid as _u
    d = request.get_json()
    max_ord = q('SELECT MAX(sort_order) as m FROM networth_assets WHERE user_id=?', (uid(),), fetchone=True)
    nxt = (max_ord['m'] or 0) + 1 if max_ord else 1
    q('INSERT INTO networth_assets (id,user_id,account,asset_class,label,amount,sort_order) VALUES (?,?,?,?,?,?,?)',
      (str(_u.uuid4()), uid(), d['account'], d['asset_class'], d.get('label',''), float(d['amount']), nxt))
    return jsonify({'ok': True})

@app.route('/api/nw/assets/<aid>', methods=['PUT'])
@login_required
def nw_update_asset(aid):
    d = request.get_json()
    q('UPDATE networth_assets SET account=?,asset_class=?,label=?,amount=? WHERE id=? AND user_id=?',
      (d['account'], d['asset_class'], d.get('label',''), float(d['amount']), aid, uid()))
    return jsonify({'ok': True})

@app.route('/api/nw/assets/<aid>', methods=['DELETE'])
@login_required
def nw_delete_asset(aid):
    q('DELETE FROM networth_assets WHERE id=? AND user_id=?', (aid, uid()))
    return jsonify({'ok': True})

@app.route('/api/nw/plan', methods=['GET'])
@login_required
def nw_get_plan():
    import json as _j
    row = q('SELECT * FROM networth_plan WHERE user_id=? ORDER BY updated_at DESC LIMIT 1', (uid(),), fetchone=True)
    if not row:
        return jsonify({'current_age':30,'current_nw':5000000,'roi_pct':15,
                        'swp_pct':3,'swp_start_age':55,'annual_invest':{}})
    row['annual_invest'] = _j.loads(row.get('annual_invest_json') or '{}')
    row['invest_rows']   = _j.loads(row.get('invest_rows_json') or '[]')
    return jsonify(row)

@app.route('/api/nw/plan', methods=['POST'])
@login_required
def nw_save_plan():
    import json as _j
    import uuid as _u2
    d = request.get_json()
    q('DELETE FROM networth_plan WHERE user_id=?', (uid(),))
    q("""INSERT INTO networth_plan
         (id,user_id,current_age,current_nw,roi_pct,swp_pct,swp_start_age,annual_invest_json,invest_rows_json,updated_at)
         VALUES (?,?,?,?,?,?,?,?,?,datetime('now'))""",
      (str(_u2.uuid4()), uid(),
       int(d.get('current_age',30)), float(d.get('current_nw',5000000)),
       float(d.get('roi_pct',15)), float(d.get('swp_pct',3)),
       int(d.get('swp_start_age',55)),
       _j.dumps(d.get('annual_invest',{})),
       _j.dumps(d.get('invest_rows',[]))))
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# AUTO BASKET MODULE API
# ══════════════════════════════════════════════════════════════════════════════

import json as _ab_json, uuid as _ab_uuid
from datetime import datetime as _dt

def ab_kite_price(instruments, access_token, api_key):
    """Fetch LTP for basket instruments.
    Strategy:
      1. Pull /portfolio/holdings  (already authenticated, returns last_price reliably)
      2. For any symbols NOT in holdings, call /quote/ltp as fallback
    Returns: { 'TRADINGSYMBOL_UPPER': ltp_float, ... }
    """
    import urllib.parse as _uparse
    if not instruments: return {}

    needed = {}   # UPPER_SYM -> exchange
    for ins in instruments:
        if isinstance(ins, dict):
            sym  = ins['tradingsymbol'].upper()
            exch = (ins.get('exchange') or 'NSE').upper()
        else:
            sym  = str(ins).upper()
            exch = 'NSE'
        needed[sym] = exch

    out = {}

    # ── Source 1: portfolio holdings (proven, returns last_price) ──────────────
    hold_resp = kite_api('/portfolio/holdings', access_token, api_key)
    if hold_resp.get('status') == 'success':
        for h in hold_resp.get('data', []):
            sym = h.get('tradingsymbol', '').upper()
            ltp = float(h.get('last_price', 0))
            if sym in needed and ltp > 0:
                out[sym] = ltp

    # ── Source 2: /quote/ltp for anything still missing ───────────────────────
    missing = [sym for sym in needed if sym not in out]
    if missing:
        pairs = [f"{needed[sym]}:{sym}" for sym in missing]
        # Kite expects repeated 'i' params: /quote/ltp?i=NSE:X&i=NSE:Y
        qs = '&'.join(f'i={_uparse.quote(p)}' for p in pairs)
        resp = kite_api(f'/quote/ltp?{qs}', access_token, api_key)
        if resp.get('status') == 'success':
            for k, v in resp.get('data', {}).items():
                sym = k.split(':', 1)[-1].upper()
                ltp = float(v.get('last_price', 0))
                if ltp > 0:
                    out[sym] = ltp

    return out


def ab_nse_price(symbols):
    """Fallback: fetch LTP from NSE public JSON (no auth needed).
    Returns: { 'TRADINGSYMBOL_UPPER': ltp_float, ... }
    """
    import urllib.request as _ur
    import json as _json
    out = {}
    for sym in symbols:
        try:
            url = f'https://www.nseindia.com/api/quote-equity?symbol={_ur.quote(sym)}'
            req = _ur.Request(url)
            req.add_header('User-Agent', 'Mozilla/5.0')
            req.add_header('Accept', 'application/json')
            req.add_header('Referer', 'https://www.nseindia.com')
            with _ur.urlopen(req, timeout=8) as r:
                data = _json.loads(r.read().decode())
            ltp = float(data.get('priceInfo', {}).get('lastPrice', 0))
            if ltp > 0:
                out[sym.upper()] = ltp
        except Exception:
            pass
    return out

@app.route('/api/ab/baskets', methods=['GET'])
@login_required
def ab_get_baskets():
    baskets = q('SELECT * FROM ab_baskets WHERE user_id=? ORDER BY created_at', (uid(),), fetchall=True)
    for b in baskets:
        instrs = q('SELECT * FROM ab_instruments WHERE basket_id=? ORDER BY sort_order',
                   (b['id'],), fetchall=True)
        b['instruments'] = instrs
        # Latest NAV
        nav_row = q('SELECT nav, total_value, nav_date FROM ab_nav_history WHERE basket_id=? ORDER BY nav_date DESC LIMIT 1',
                    (b['id'],), fetchone=True)
        if nav_row:
            b['nav']          = nav_row['nav']
            b['latest_value'] = nav_row['total_value']
            b['nav_date']     = nav_row['nav_date']
        b['return_pct'] = ((b['nav'] - 100.0) / 100.0 * 100) if b['nav'] else 0
    return jsonify(baskets)

@app.route('/api/ab/baskets', methods=['POST'])
@login_required
def ab_create_basket():
    d = request.get_json()
    bid = str(_ab_uuid.uuid4())
    inception = d.get('inception') or _dt.now().strftime('%Y-%m-%d')
    q('''INSERT INTO ab_baskets (id,user_id,name,strategy,rebalance,capital,nav,inception,status,notes)
         VALUES (?,?,?,?,?,?,100.0,?,?,?)''',
      (bid, uid(), d['name'], d.get('strategy',''), d.get('rebalance','monthly'),
       float(d.get('capital',0)), inception, 'active', d.get('notes','')))
    # Save instruments
    for i, ins in enumerate(d.get('instruments', [])):
        iid = str(_ab_uuid.uuid4())
        q('''INSERT INTO ab_instruments (id,basket_id,tradingsymbol,exchange,target_pct,qty,avg_price,sort_order)
             VALUES (?,?,?,?,?,?,?,?)''',
          (iid, bid, ins['tradingsymbol'].upper(), ins.get('exchange','NSE'),
           float(ins.get('target_pct',0)), int(ins.get('qty',0)),
           float(ins.get('avg_price',0)), i))
    # Seed initial NAV history
    q('''INSERT INTO ab_nav_history (id,basket_id,nav_date,nav,total_value) VALUES (?,?,?,100.0,?)''',
      (str(_ab_uuid.uuid4()), bid, inception, float(d.get('capital',0))))
    return jsonify({'ok': True, 'id': bid})

@app.route('/api/ab/baskets/<bid>', methods=['PUT'])
@login_required
def ab_update_basket(bid):
    d = request.get_json()
    q('''UPDATE ab_baskets SET name=?,strategy=?,rebalance=?,capital=?,status=?,notes=? WHERE id=? AND user_id=?''',
      (d['name'], d.get('strategy',''), d.get('rebalance','monthly'),
       float(d.get('capital',0)), d.get('status','active'), d.get('notes',''), bid, uid()))
    # Replace instruments
    q('DELETE FROM ab_instruments WHERE basket_id=?', (bid,))
    for i, ins in enumerate(d.get('instruments', [])):
        iid = str(_ab_uuid.uuid4())
        q('''INSERT INTO ab_instruments (id,basket_id,tradingsymbol,exchange,target_pct,qty,avg_price,sort_order)
             VALUES (?,?,?,?,?,?,?,?)''',
          (iid, bid, ins['tradingsymbol'].upper(), ins.get('exchange','NSE'),
           float(ins.get('target_pct',0)), int(ins.get('qty',0)),
           float(ins.get('avg_price',0)), i))
    return jsonify({'ok': True})

@app.route('/api/ab/baskets/<bid>', methods=['DELETE'])
@login_required
def ab_delete_basket(bid):
    q('DELETE FROM ab_instruments WHERE basket_id=?', (bid,))
    q('DELETE FROM ab_nav_history WHERE basket_id=?', (bid,))
    q('DELETE FROM ab_orders WHERE basket_id=?', (bid,))
    q('DELETE FROM ab_baskets WHERE id=? AND user_id=?', (bid, uid()))
    return jsonify({'ok': True})

@app.route('/api/ab/baskets/<bid>/nav', methods=['GET'])
@login_required
def ab_nav_history(bid):
    rows = q('''SELECT nav_date, nav, total_value FROM ab_nav_history
                WHERE basket_id=? GROUP BY nav_date ORDER BY nav_date''',
             (bid,), fetchall=True)
    return jsonify(rows)

@app.route('/api/ab/baskets/<bid>/refresh', methods=['POST'])
@login_required
def ab_refresh_nav(bid):
    """Fetch live prices, recompute NAV and record it."""
    access_token = kite_cfg_get('access_token')
    api_key      = kite_cfg_get('api_key')
    basket = q('SELECT * FROM ab_baskets WHERE id=? AND user_id=?', (bid, uid()), fetchone=True)
    if not basket:
        return jsonify({'error': 'Basket not found'}), 404
    instrs = q('SELECT * FROM ab_instruments WHERE basket_id=?', (bid,), fetchall=True)
    if not instrs:
        return jsonify({'error': 'No instruments'}), 400

    prices = {}
    if access_token and api_key:
        prices = ab_kite_price(instrs, access_token, api_key)

    # NSE fallback for any symbols still missing after Kite fetch
    missing_syms = [i['tradingsymbol'].upper() for i in instrs if not prices.get(i['tradingsymbol'].upper())]
    if missing_syms:
        nse_prices = ab_nse_price(missing_syms)
        prices.update(nse_prices)

    total_value = 0.0
    updated = []
    effective_prices = {}  # keyed by UPPERCASE tradingsymbol for consistent lookup
    for ins in instrs:
        sym_upper = ins['tradingsymbol'].upper()
        ltp = prices.get(sym_upper, 0)
        if not ltp:
            ltp = float(ins.get('avg_price', 0))
        effective_prices[sym_upper] = ltp  # always store uppercase
        val = ltp * int(ins['qty'])
        total_value += val
        updated.append({**ins, 'ltp': ltp, 'value': val,
                        'pnl': (ltp - ins['avg_price']) * ins['qty'],
                        'pnl_pct': ((ltp - ins['avg_price']) / ins['avg_price'] * 100) if ins['avg_price'] else 0})

    # Compute NAV = 100 × (current_value / inception_capital)
    capital = float(basket['capital']) or 1
    nav = 100.0 * total_value / capital if total_value > 0 else basket['nav']

    today = _dt.now().strftime('%Y-%m-%d')
    # Upsert nav history for today
    existing = q('SELECT id FROM ab_nav_history WHERE basket_id=? AND nav_date=?',
                 (bid, today), fetchone=True)
    if existing:
        q('UPDATE ab_nav_history SET nav=?,total_value=? WHERE basket_id=? AND nav_date=?',
          (round(nav,4), round(total_value,2), bid, today))
    else:
        q('INSERT INTO ab_nav_history (id,basket_id,nav_date,nav,total_value) VALUES (?,?,?,?,?)',
          (str(_ab_uuid.uuid4()), bid, today, round(nav,4), round(total_value,2)))

    q('UPDATE ab_baskets SET nav=? WHERE id=?', (round(nav,4), bid))

    return jsonify({'ok': True, 'nav': round(nav,4), 'total_value': round(total_value,2),
                    'instruments': updated, 'prices': effective_prices})

@app.route('/api/ab/baskets/<bid>/rebalance', methods=['POST'])
@login_required
def ab_rebalance(bid):
    """Calculate rebalance orders. place=true to actually place via Kite."""
    import math
    access_token = kite_cfg_get('access_token')
    api_key      = kite_cfg_get('api_key')
    d = request.get_json() or {}
    place = d.get('place', False)

    basket = q('SELECT * FROM ab_baskets WHERE id=? AND user_id=?', (bid, uid()), fetchone=True)
    if not basket: return jsonify({'error': 'Basket not found'}), 404
    instrs = q('SELECT * FROM ab_instruments WHERE basket_id=?', (bid,), fetchall=True)
    if not instrs: return jsonify({'error': 'No instruments'}), 400

    prices = {}
    if access_token and api_key:
        prices = ab_kite_price(instrs, access_token, api_key)

    capital = float(basket['capital'])
    orders = []
    for ins in instrs:
        ltp = prices.get(ins['tradingsymbol'].upper(), ins.get('avg_price', 0))
        if ltp <= 0: continue
        target_val = capital * float(ins['target_pct']) / 100.0
        target_qty = int(target_val / ltp)
        current_qty = int(ins['qty'])
        diff = target_qty - current_qty
        if diff == 0: continue
        txn = 'BUY' if diff > 0 else 'SELL'
        order = {'tradingsymbol': ins['tradingsymbol'], 'exchange': ins['exchange'],
                 'transaction_type': txn, 'qty': abs(diff),
                 'price': round(ltp, 2), 'target_qty': target_qty,
                 'current_qty': current_qty, 'value': round(abs(diff) * ltp, 2)}
        orders.append(order)

        if place and access_token and api_key:
            # Place market order via Kite
            order_data = _ab_json.dumps({
                'tradingsymbol': ins['tradingsymbol'], 'exchange': ins['exchange'],
                'transaction_type': txn, 'order_type': 'MARKET',
                'product': 'CNC', 'quantity': abs(diff), 'validity': 'DAY'
            }).encode()
            import urllib.request as _ur
            req = _ur.Request('https://api.kite.trade/orders/regular', data=order_data, method='POST')
            req.add_header('X-Kite-Version', '3')
            req.add_header('Authorization', f'token {api_key}:{access_token}')
            req.add_header('Content-Type', 'application/json')
            try:
                with _ur.urlopen(req, timeout=15) as resp:
                    result = _ab_json.loads(resp.read().decode())
                kite_order_id = result.get('data', {}).get('order_id', '')
                status = 'placed'
            except Exception as ex:
                kite_order_id = ''
                status = f'failed: {ex}'
            # Record order
            q('''INSERT INTO ab_orders (id,basket_id,order_id,tradingsymbol,exchange,transaction_type,qty,price,status,reason)
                 VALUES (?,?,?,?,?,?,?,?,?,?)''',
              (str(_ab_uuid.uuid4()), bid, kite_order_id, ins['tradingsymbol'],
               ins['exchange'], txn, abs(diff), round(ltp,2), status, 'rebalance'))

    return jsonify({'ok': True, 'orders': orders, 'placed': place})

@app.route('/api/ab/baskets/<bid>/orders', methods=['GET'])
@login_required
def ab_get_orders(bid):
    rows = q('SELECT * FROM ab_orders WHERE basket_id=? ORDER BY created_at DESC',
             (bid,), fetchall=True)
    return jsonify(rows)

@app.route('/api/ab/nifty', methods=['GET'])
@login_required
def ab_nifty():
    """Fetch live Nifty 50 price — tries Kite first, then NSE public API."""
    import urllib.request as _ur, urllib.parse as _up, json as _j

    # ── Try Kite first (indices use NSE:NIFTY 50 instrument token) ─────────────
    access_token = kite_cfg_get('access_token')
    api_key      = kite_cfg_get('api_key')
    if access_token and api_key:
        # Correct Kite instrument key for Nifty index is "NSE:NIFTY 50"
        qs = 'i=' + _up.quote('NSE:NIFTY 50')
        resp = kite_api(f'/quote/ltp?{qs}', access_token, api_key)
        if resp.get('status') == 'success':
            # Kite may return key as "NSE:NIFTY 50" or "NSE:NIFTY50"
            data = resp.get('data', {})
            for k, v in data.items():
                ltp = float(v.get('last_price', 0))
                if ltp > 0:
                    return jsonify({'ok': True, 'ltp': ltp, 'source': 'kite'})

    # ── Fallback: NSE public API (no auth) ─────────────────────────────────────
    try:
        req = _ur.Request('https://www.nseindia.com/api/allIndices')
        req.add_header('User-Agent', 'Mozilla/5.0')
        req.add_header('Accept', 'application/json')
        req.add_header('Referer', 'https://www.nseindia.com')
        with _ur.urlopen(req, timeout=8) as r:
            indices = _j.loads(r.read().decode()).get('data', [])
        for idx in indices:
            if idx.get('index') in ('NIFTY 50', 'Nifty 50', 'NIFTY50'):
                ltp = float(idx.get('last', 0))
                if ltp > 0:
                    return jsonify({'ok': True, 'ltp': ltp, 'source': 'nse'})
    except Exception as e:
        pass

    return jsonify({'error': 'Could not fetch Nifty price from Kite or NSE'}), 200

@app.route('/api/ab/summary', methods=['GET'])
@login_required
def ab_summary():
    baskets = q("SELECT * FROM ab_baskets WHERE user_id=? AND status='active'", (uid(),), fetchall=True)
    total_capital = sum(float(b['capital']) for b in baskets)
    total_nav_val = sum(float(b.get('nav',100)) * float(b['capital']) / 100 for b in baskets)
    total_pnl = total_nav_val - total_capital

    # Per-basket allocation % based on invested capital
    basket_list = []
    for b in baskets:
        cap = float(b['capital'])
        alloc_pct = round((cap / total_capital * 100), 2) if total_capital > 0 else 0
        basket_list.append({'id': b['id'], 'name': b['name'], 'capital': cap,
                            'alloc_pct': alloc_pct})

    return jsonify({
        'basket_count': len(baskets),
        'total_capital': total_capital,
        'total_value': total_nav_val,
        'total_pnl': total_pnl,
        'baskets': basket_list,
    })

@app.route('/api/ab/debug_prices', methods=['GET'])
@login_required
def ab_debug_prices():
    """Debug: show prices from all sources for all basket instruments."""
    import urllib.parse as _uparse
    access_token = kite_cfg_get('access_token')
    api_key      = kite_cfg_get('api_key')
    baskets = q("SELECT * FROM ab_baskets WHERE user_id=? AND status='active'", (uid(),), fetchall=True)
    all_instrs = []
    for b in baskets:
        rows = q('SELECT tradingsymbol, exchange, avg_price FROM ab_instruments WHERE basket_id=?', (b['id'],), fetchall=True)
        all_instrs.extend(rows)

    result = {'instruments': [i['tradingsymbol'] for i in all_instrs],
              'kite_connected': bool(access_token and api_key)}

    # Raw Kite /portfolio/holdings (source 1)
    if access_token and api_key:
        hold_resp = kite_api('/portfolio/holdings', access_token, api_key)
        holdings_prices = {}
        if hold_resp.get('status') == 'success':
            for h in hold_resp.get('data', []):
                holdings_prices[h.get('tradingsymbol','').upper()] = float(h.get('last_price',0))
        result['kite_holdings_prices'] = holdings_prices

        # Raw Kite /quote/ltp (source 2)
        pairs = [f"{(i.get('exchange') or 'NSE').upper()}:{i['tradingsymbol'].upper()}" for i in all_instrs]
        qs = '&'.join(f'i={_uparse.quote(p)}' for p in pairs)
        raw_ltp = kite_api(f'/quote/ltp?{qs}', access_token, api_key)
        result['kite_quote_ltp_query'] = qs
        result['kite_quote_ltp_response'] = raw_ltp

    # Final merged prices
    result['final_prices'] = ab_kite_price(all_instrs, access_token, api_key) if (access_token and api_key) else {}

    # Per-instrument verdict
    verdict = []
    for ins in all_instrs:
        sym = ins['tradingsymbol'].upper()
        ltp = result['final_prices'].get(sym, 0)
        verdict.append({'symbol': sym, 'exchange': ins.get('exchange','NSE'),
                        'avg_price': ins['avg_price'], 'ltp_fetched': ltp,
                        'source': 'live' if ltp and ltp != ins['avg_price'] else 'avg_fallback'})
    result['verdict'] = verdict
    return jsonify(result)

@app.route('/api/ab/holdings_perf', methods=['GET'])
@login_required
def ab_holdings_perf():
    """Top/bottom performers across all baskets by fetching live prices."""
    access_token = kite_cfg_get('access_token')
    api_key      = kite_cfg_get('api_key')
    baskets = q("SELECT * FROM ab_baskets WHERE user_id=? AND status='active'", (uid(),), fetchall=True)
    all_instrs = []
    basket_map = {b['id']: b['name'] for b in baskets}
    for b in baskets:
        rows = q('SELECT * FROM ab_instruments WHERE basket_id=?', (b['id'],), fetchall=True)
        for r in rows:
            r['basket_name'] = b['name']
            all_instrs.append(r)

    # Deduplicate instruments by (exchange, tradingsymbol)
    seen = set(); uniq_instrs = []
    for i in all_instrs:
        key = (i.get('exchange','NSE'), i['tradingsymbol'])
        if key not in seen:
            seen.add(key); uniq_instrs.append(i)
    prices = {}
    if access_token and api_key and uniq_instrs:
        prices = ab_kite_price(uniq_instrs, access_token, api_key)

    perf = []
    for ins in all_instrs:
        ltp = prices.get(ins['tradingsymbol'].upper(), 0)
        if not ltp:
            ltp = float(ins.get('avg_price', 0))  # fallback to avg_price when Kite price unavailable
        avg = float(ins.get('avg_price', 0))
        pnl_pct = ((ltp - avg) / avg * 100) if avg else 0
        day_chg = 0  # would need prev close for accurate day change
        perf.append({'tradingsymbol': ins['tradingsymbol'],
                     'basket_name': ins['basket_name'],
                     'ltp': ltp, 'avg': avg,
                     'pnl_pct': round(pnl_pct, 2),
                     'qty': ins['qty']})

    perf.sort(key=lambda x: x['pnl_pct'], reverse=True)
    return jsonify({'instruments': perf})

# ── Account info API ─────────────────────────────────────────────────────────
@app.route('/api/me')
@login_required
def api_me():
    user = q('SELECT id,username,email,is_admin,created_at FROM users WHERE id=?',
             (uid(),), fetchone=True)
    return jsonify(user or {})

if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    init_db()
    init_kite_db()
    port = int(os.environ.get('PORT', 5000))
    debug = not os.environ.get('RENDER', False)
    host = '0.0.0.0' if os.environ.get('RENDER') else '127.0.0.1'
    print(f"🚀 FinHub running at http://{host}:{port}")
    print("👥 Multi-user mode — /register to create accounts")
    ab_routes = [r for r in app.url_map.iter_rules() if '/ab/' in r.rule]
    print(f"🧺 AutoBasket: {len(ab_routes)} routes registered" if ab_routes else "❌ AutoBasket routes NOT found!")
    app.run(debug=debug, host=host, port=port)
